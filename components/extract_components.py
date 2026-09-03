"""Extract simple, single-table component lists from RDK-B_Component_List_2026.xlsx.

Repo identity (name, category, URL, CORE flag) comes from the 'Components'
sheet -- one clean row per repo, versus 'All Profiles'' much messier
per-feature breakdown where the same real repo often appears under several
different spelling variants (e.g. "Utopia" / "utopia" / "utopia/P&M" /
"provisioning-and-management,Utopia" -- five variants for what looks like
one or two real repos). Grouping by exact string match against that sheet
would split single components into multiple near-duplicate entries, so repo
identity stays anchored to 'Components'.

Required/Optional/n/a *classification* per profile comes from 'All Profiles'
whenever the repo's name matches there, using a normalized comparison
(lowercase, punctuation/spacing stripped -- e.g. "WebUI" matches "webui",
"WAN Manager" matches "wan-manager") rather than a brittle exact-string
match, since spot-checks (and a full cross-sheet diff) found both
confirmed-stale values in 'Components' (e.g. "Mesh Agent" marked n/a there
while both 'All Profiles' and 'Router' agree it should be
Optional/Required) and real repos the old exact-match missed purely over
case/punctuation (e.g. "WebUI" vs "webui" -- same repo, All Profiles says
n/a, but exact matching silently failed to apply that). A repo counts as
Required/Optional/n/a in 'All Profiles' based on the best (most permissive:
Required beats Optional beats n/a) status seen across all rows there that
normalize to the same key. Repos whose name has no normalized match at all
in 'All Profiles' (e.g. "dhcp-manager (recipe for DHCP client only)", a
Components-only naming variant with no counterpart of any spelling there)
keep 'Components'' original classification, since there's nothing reliable
to look up.

Three modes:

  core         -> components tagged CORE (common to every profile) plus components
                  that are Required in every profile they apply to. Tiers:
                  'common-core' vs 'required'.

  profile      -> all components that apply (Required or Optional, i.e. not 'n/a')
                  to a single device profile column, e.g. "EthWAN WiFi Router".
                  Tiers: 'required' vs 'optional'.

  all-profiles -> runs 'profile' for every column in PROFILE_COLUMNS in one go,
                  writing one <profile>-components.json per profile into the
                  current directory (see PROFILE_FILENAMES for exact names).

  all-components -> every component relevant to RDK-B Core Broadband across
                  ANY profile -- Common Core, Required, or Optional for at
                  least one profile. Not scoped to a single profile; fits
                  profile-agnostic pages (e.g. north-bound-apis.html, since
                  the North Bound API surface isn't tied to one device
                  profile). Tiers: 'common-core' / 'required' / 'optional',
                  a single collapsed summary per component -- no
                  per-profile breakdown.

Usage:
    python3 extract_components.py core \
        --xlsx RDK-B_Component_List_2026.xlsx --out core-b-components.json

    python3 extract_components.py profile "EthWAN WiFi Router" \
        --xlsx RDK-B_Component_List_2026.xlsx --out ethwan-router-components.json

    python3 extract_components.py all-profiles \
        --xlsx RDK-B_Component_List_2026.xlsx --show-core
"""
from __future__ import annotations

import argparse
import json
import re
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

SCHEMA_VERSION = "1.0"

PROFILE_COLUMNS = [
    "Modem\n/ONU",
    "EthWAN WiFi Router",
    "GW",
    "GW\nOpenSync",
    "GW\nEasyMesh",
    "EXT\nOpenSync",
    "EXT\nEasyMesh",
]

# Output filename per profile column, for the "all-profiles" mode.
# EthWAN WiFi Router keeps its existing name since components/index.html and
# components/gen_components_page.py already depend on it by that exact
# filename; every other profile follows <profile-id>-components.json.
PROFILE_FILENAMES = {
    "Modem\n/ONU": "modem-onu-components.json",
    "EthWAN WiFi Router": "ethwan-router-components.json",
    "GW": "gw-components.json",
    "GW\nOpenSync": "gw-opensync-components.json",
    "GW\nEasyMesh": "gw-easymesh-components.json",
    "EXT\nOpenSync": "ext-opensync-components.json",
    "EXT\nEasyMesh": "ext-easymesh-components.json",
}

TIERS = {
    "common-core": {"id": "common-core", "label": "Common Core", "color": "gold"},
    "required": {"id": "required", "label": "Required", "color": "blue"},
    "optional": {"id": "optional", "label": "Optional", "color": "gray"},
}


def _norm_key(name: str) -> str:
    """Lowercase, alphanumeric-only key for matching the same real repo
    across the two sheets' inconsistent naming (case, spaces vs hyphens vs
    underscores, punctuation -- e.g. "WebUI" / "webui", "WAN Manager" /
    "wan-manager", "T2 Telemetry" / "T2 telemetry"). Verified against the
    full repo list: zero collisions on the 'Components' side (the one that
    matters for lookup safety here); the couple of collisions this creates
    within 'All Profiles' itself (e.g. "dcm agent" / "dcm-agent") are
    confirmed re-entries of the same real repo, not different repos, so
    aggregating them together is correct."""
    return re.sub(r"[^a-z0-9]", "", name.lower())


# A handful of repos are confirmed to be the same real component across the
# two sheets, but named too differently for _norm_key's punctuation/case
# normalization to catch -- verified by matching each pair's Component
# Description (Components sheet) against its Detailed Feature List text
# (All Profiles sheet), not guessed from name similarity alone:
#   - "USP PA" / "usp-pa-vendor-rdk": both describe USP (TR-369) device
#     management.
#   - "Web PA including: ..." / the parodus+start-parodus+xmidt-org bundle
#     row: both describe WebPA cloud device management over WebSockets.
#   - "Break pad" / "breakpad_wrapper": Components' own description is
#     literally "A wrapper for breakpad features".
#   - "dhcp-manager (recipe for DHCP client only)" / generic "dhcp-manager":
#     Components' own Notes text ties this variant to the WAN/LAN client
#     capability the generic All-Profiles rows cover.
# Maps a Components-sheet repo name to the normalized key to look up in
# 'All Profiles' instead of the repo's own (which has no match there).
CURATED_ALIASES: dict[str, str] = {
    "USP PA": _norm_key("usp-pa-vendor-rdk"),
    "Web PA\n\nincluding:\n- parodus\n- parodus2ccsp\n- start-parodus\n- wdmp-c": _norm_key("parodus\nstart-parodus\nxmidt-org"),
    "Break pad": _norm_key("breakpad_wrapper"),
    "dhcp-manager (recipe for DHCP client only)": _norm_key("dhcp-manager"),
}


def _all_profiles_classification(wb) -> dict[str, dict[str, str]]:
    """Aggregate the 'All Profiles' sheet into {normalized_repo_name:
    {profile: status}}, one status per (repo, profile) -- the best (most
    permissive) status seen across that repo's several detail-feature rows
    there, AND across any other repo name in 'All Profiles' that normalizes
    to the same key (see _norm_key). Used only to correct
    'Components'-sourced classifications where they disagree, not as a
    source of repo identity (see module docstring)."""
    ws = wb["All Profiles"]
    headers = [c.value for c in ws[1]]
    name_idx = headers.index("Component Repo")
    profile_idxs = [headers.index(c) for c in PROFILE_COLUMNS]
    RANK = {"Required": 2, "Optional": 1, "n/a": 0}

    best: dict[str, dict[str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[name_idx]
        if name is None:
            continue
        if isinstance(name, str):
            name = name.strip()
        if not name:
            continue
        key = _norm_key(name)
        per_profile = best.setdefault(key, {})
        for i, profile in enumerate(PROFILE_COLUMNS):
            v = row[profile_idxs[i]]
            if v not in RANK:
                continue
            cur = per_profile.get(profile)
            if cur is None or RANK[v] > RANK[cur]:
                per_profile[profile] = v
    return best


def _split_urls(cell) -> list[str]:
    """Split a Github Link cell into one or more URLs. Most cells hold a
    single URL; a handful list several -- one per line, sometimes comma-
    or semicolon-separated on one line (see module docstring). Any
    resulting fragment that doesn't look like a URL is dropped rather than
    surfaced as a broken link -- some of these cells mix in plain-text
    notes alongside the real links."""
    if not isinstance(cell, str):
        return []
    parts = re.split(r"[\n,;]+", cell)
    return [p.strip() for p in parts if p.strip().lower().startswith("http")]


def _rows(wb):
    """Yield one record per repo from the 'Components' sheet (clean, one row
    per repo -- see module docstring for why repo identity stays anchored
    here rather than to 'All Profiles'). Required/Optional/n/a values come
    from 'All Profiles' wherever the repo's exact name is found there;
    repos absent from 'All Profiles' by that exact name keep their original
    'Components' classification unchanged."""
    ws = wb["Components"]
    corrections = _all_profiles_classification(wb)

    headers = [c.value for c in ws[1]]
    name_idx = headers.index("Component Repo")
    subsys_idx = headers.index("Subsystem")
    url_idx = headers.index("Github Link")
    core_idx = headers.index("Core Components")
    profile_idxs = [headers.index(c) for c in PROFILE_COLUMNS]

    cur_subsys = None
    seen_names = set()
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[name_idx] is None and row[subsys_idx] is None:
            continue
        if row[subsys_idx]:
            cur_subsys = row[subsys_idx]
        name = row[name_idx]
        if name is None:
            continue
        if isinstance(name, str):
            name = name.strip()
        # A few repos are accidentally listed twice as separate rows in the
        # source sheet (e.g. "rdk-cert-config" at rows 77 and 78, identical
        # classification, one missing its Subsystem) -- keep the first
        # occurrence only.
        if name in seen_names:
            continue
        seen_names.add(name)
        urls = _split_urls(row[url_idx])
        url = urls[0] if urls else None
        supporting_urls = urls[1:]

        profile_values = {}
        for i, profile in enumerate(PROFILE_COLUMNS):
            orig = row[profile_idxs[i]]
            lookup_key = CURATED_ALIASES.get(name, _norm_key(name))
            corrected = corrections.get(lookup_key, {}).get(profile)
            # All Profiles' classification wins outright whenever the repo
            # name matches there, not only when it disagrees on inclusion
            # (Required/Optional vs n/a) -- it also settles the finer-grained
            # Required-vs-Optional question the same way, e.g. WebConfig
            # (Components: Optional, All Profiles: Required for GW OpenSync).
            profile_values[profile] = corrected if corrected is not None else orig

        yield {
            "subsystem": cur_subsys,
            "name": name,
            "url": url,
            "supporting_urls": supporting_urls,
            "is_core": row[core_idx] == "CORE",
            "profile_values": profile_values,
        }


def extract_all(wb) -> list[dict]:
    """Every component in 'Components' with any real classification anywhere
    (Required, Optional, or CORE for at least one profile) -- the full
    RDK-B component universe, not scoped to one device profile. Fits pages
    like north-bound-apis.html where the subject (protocol/API surface) is
    profile-agnostic by nature: a component can be n/a for one profile and
    Optional for another, and this list doesn't care which -- it's in if
    it's relevant to RDK-B Core Broadband anywhere.

    tier is a single collapsed summary, most-significant first:
    'common-core' (CORE-flagged) > 'required' (Required for at least one
    profile) > 'optional' (Optional for at least one profile, never
    Required or CORE). No per-profile breakdown -- see module/CLI docs if
    that's ever needed instead.
    """
    out = []
    for r in _rows(wb):
        vals = [v for v in r["profile_values"].values() if v not in (None, "n/a")]
        if r["is_core"]:
            tier = "common-core"
        elif "Required" in vals:
            tier = "required"
        elif "Optional" in vals:
            tier = "optional"
        else:
            continue  # n/a everywhere -- not part of the RDK-B component universe
        out.append({
            "name": r["name"],
            "category": r["subsystem"],
            "tier": tier,
            "url": r["url"],
            "supportingUrls": r["supporting_urls"],
        })
    return out


def extract_core(wb) -> list[dict]:
    out = []
    for r in _rows(wb):
        vals = [v for v in r["profile_values"].values() if v not in (None, "n/a")]
        required_everywhere = bool(vals) and all(v == "Required" for v in vals)
        if r["is_core"] or required_everywhere:
            out.append({
                "name": r["name"],
                "category": r["subsystem"],
                "tier": "common-core" if r["is_core"] else "required",
                "url": r["url"],
                "supportingUrls": r["supporting_urls"],
            })
    return out


def extract_profile(wb, profile: str, required_only: bool = False, show_core: bool = False) -> list[dict]:
    """All components that apply to a single device profile column.

    show_core=False (default): tiers are 'required' / 'optional' only, matching
        the profile column value verbatim (today's simple pages).
    show_core=True: components flagged CORE ('Core Components' == 'CORE') are
        tagged 'common-core' instead of their raw Required/Optional value, so
        the page distinguishes Common Core vs Required vs Optional. Ignored
        when required_only=True (core components are always required-everywhere
        by definition, so they'd show as Required either way).
    """
    if profile not in PROFILE_COLUMNS:
        raise SystemExit(f"Unknown profile {profile!r}. Choose from: {PROFILE_COLUMNS}")
    wanted = ("Required",) if required_only else ("Required", "Optional")
    out = []
    for r in _rows(wb):
        v = r["profile_values"][profile]
        if v not in wanted:
            continue
        tier = "common-core" if (show_core and not required_only and r["is_core"]) else v.lower()
        out.append({
            "name": r["name"],
            "category": r["subsystem"],
            "tier": tier,
            "url": r["url"],
            "supportingUrls": r["supporting_urls"],
        })
    return out


def build_payload(components: list[dict], title: str, subtitle: str, tier_ids: list[str]) -> dict:
    return {
        "schemaVersion": SCHEMA_VERSION,
        "generatedAt": datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z"),
        "title": title,
        "subtitle": subtitle,
        "tiers": [TIERS[t] for t in tier_ids],
        "components": components,
    }


def build_profile_payload(wb, profile: str, required_only: bool = False, show_core: bool = False) -> tuple[list[dict], dict]:
    """Extract + build the payload for one profile column. Shared by the
    'profile' and 'all-profiles' CLI modes so their output is identical."""
    components = extract_profile(wb, profile, required_only=required_only, show_core=show_core)
    if required_only:
        tier_ids = ["required"]
    elif show_core:
        tier_ids = ["common-core", "required", "optional"]
    else:
        tier_ids = ["required", "optional"]
    # Some profile column headers in the xlsx contain an embedded newline
    # (e.g. "GW\nOpenSync", wrapped for column width) -- that's needed
    # verbatim for the exact-match lookup in extract_profile(), but reads as
    # a broken mid-sentence line break in a title/subtitle string, so
    # normalize to a single space for display purposes only.
    profile_display = " ".join(profile.split())
    subtitle = f"Components required for the {profile_display} device profile." if required_only \
        else f"Components for the {profile_display} device profile: Common Core, Required, and Optional." if show_core \
        else f"Components that apply to the {profile_display} device profile."
    payload = build_payload(
        components,
        title=f"RDK-B {profile_display} Components",
        subtitle=subtitle,
        tier_ids=tier_ids,
    )
    return components, payload


def main() -> None:
    p = argparse.ArgumentParser()
    sub = p.add_subparsers(dest="mode", required=True)

    core_p = sub.add_parser("core")
    core_p.add_argument("--xlsx", default="RDK-B_Component_List_2026.xlsx")
    core_p.add_argument("--out", default="core-b-components.json")

    prof_p = sub.add_parser("profile")
    prof_p.add_argument("profile", help='e.g. "EthWAN WiFi Router"')
    prof_p.add_argument("--xlsx", default="RDK-B_Component_List_2026.xlsx")
    prof_p.add_argument("--out", default="profile-components.json")
    prof_p.add_argument("--required-only", action="store_true", help="Omit Optional components; show only Required.")
    prof_p.add_argument("--show-core", action="store_true", help="Tag CORE components as 'Common Core' instead of Required/Optional.")

    all_p = sub.add_parser("all-profiles", help="Run 'profile' for every profile column, writing one file each into the current directory.")
    all_p.add_argument("--xlsx", default="RDK-B_Component_List_2026.xlsx")
    all_p.add_argument("--required-only", action="store_true", help="Omit Optional components; show only Required.")
    all_p.add_argument("--show-core", action="store_true", help="Tag CORE components as 'Common Core' instead of Required/Optional.")

    allc_p = sub.add_parser("all-components", help="Every component relevant to any profile -- the full RDK-B component universe, not scoped to one profile.")
    allc_p.add_argument("--xlsx", default="RDK-B_Component_List_2026.xlsx")
    allc_p.add_argument("--out", default="all-components.json")

    args = p.parse_args()
    wb = openpyxl.load_workbook(Path(args.xlsx), data_only=True)

    if args.mode == "core":
        components = extract_core(wb)
        payload = build_payload(
            components,
            title="Core RDK-B Components",
            subtitle="Components common to every RDK-B device profile, or required wherever they apply.",
            tier_ids=["common-core", "required"],
        )
        Path(args.out).write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"Wrote {args.out} ({len(components)} components)")

    elif args.mode == "all-components":
        components = extract_all(wb)
        payload = build_payload(
            components,
            title="RDK-B Core Broadband Components",
            subtitle="Every component relevant to RDK-B Core Broadband across all device profiles -- Common Core, Required, or Optional for at least one profile.",
            tier_ids=["common-core", "required", "optional"],
        )
        Path(args.out).write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"Wrote {args.out} ({len(components)} components)")

    elif args.mode == "profile":
        components, payload = build_profile_payload(wb, args.profile, required_only=args.required_only, show_core=args.show_core)
        Path(args.out).write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"Wrote {args.out} ({len(components)} components)")

    else:  # all-profiles
        for profile in PROFILE_COLUMNS:
            out_name = PROFILE_FILENAMES[profile]
            components, payload = build_profile_payload(wb, profile, required_only=args.required_only, show_core=args.show_core)
            Path(out_name).write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
            print(f"  {out_name:35} {len(components):3} components  ({' '.join(profile.split())})")
        print(f"Wrote {len(PROFILE_COLUMNS)} profile component files.")


if __name__ == "__main__":
    main()
