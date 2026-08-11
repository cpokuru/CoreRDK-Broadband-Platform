"""Convert RDK-B_Component_List_2026.xlsx -> RDK-B_Component_List_2026.json

Produces JSON that validates against rdk-b-components.schema.json and has the
exact same {name, headers, rows} shape that the page used to get at runtime
from XLSX.utils.sheet_to_json(ws, {header: 1}) via SheetJS in the browser.
That means index.html's rendering code (initWithSheets onward) needs zero
changes -- only the loading layer changes from "parse .xlsx in the browser"
to "fetch this .json".

Usage:
    python3 xlsx_to_json.py [source.xlsx] [dest.json]
"""
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

SCHEMA_VERSION = "1.0"


def normalize_cell(v) -> str:
    """Mirror the JS normalizeCell(): stringify, normalize newlines, blank for None."""
    if v is None:
        return ""
    s = str(v)
    return s.replace("\r\n", "\n").replace("\r", "\n")


def sheet_to_dict(ws) -> dict:
    """Mirror parseWorkbookBuffer()'s per-sheet logic exactly."""
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return {"name": ws.title, "headers": [], "rows": []}

    headers = [normalize_cell(v) for v in all_rows[0]]
    rows = []
    for raw_row in all_rows[1:]:
        # Pad/truncate to header length, exactly like `headers.map((_, i) => normalizeCell(r[i]))`
        row = [
            normalize_cell(raw_row[i]) if i < len(raw_row) else ""
            for i in range(len(headers))
        ]
        if any(cell.strip() != "" for cell in row):
            rows.append(row)

    return {"name": ws.title, "headers": headers, "rows": rows}


def convert(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheets = [sheet_to_dict(wb[name]) for name in wb.sheetnames]
    return {
        "schemaVersion": SCHEMA_VERSION,
        "generatedAt": datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z"),
        "sourceWorkbook": xlsx_path.name,
        "sheets": sheets,
    }


def main() -> None:
    root = Path(__file__).parent
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else root / "RDK-B_Component_List_2026.xlsx"
    dest = Path(sys.argv[2]) if len(sys.argv) > 2 else root / "RDK-B_Component_List_2026.json"

    data = convert(src)
    dest.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")

    n_rows = sum(len(s["rows"]) for s in data["sheets"])
    print(f"Wrote {dest} ({len(data['sheets'])} sheets, {n_rows} rows total)")


if __name__ == "__main__":
    main()
